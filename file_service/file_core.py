from utils import QuizDatabase
from openpyxl import load_workbook
from file_service import get_path
from data import engine

db = QuizDatabase(engine=engine)


async def read_file(file_path: str, subject_id: int):
    try:
        sheet = load_workbook(await get_path(file_name=file_path)).active
        if await db.get_question(subject_id=subject_id, text=next(sheet.iter_rows(values_only=True))[1]):
            print("Malumot bazasida savol mavjud")
            return None
        else:
            print("Malumot bazasida savol mavjud emas")
            for row in sheet.iter_rows(values_only=True):
                await db.add_question(subject_id=subject_id, question=row[1], option1=row[2], option2=row[3], option3=row[4],
                                      option4=row[5])
    except Exception as e:
        print(e)
        return None

